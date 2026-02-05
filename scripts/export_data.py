#!/usr/bin/env python3
"""
Comprehensive Export Script for AutoCognitix

Exports data to multiple portable formats with advanced filtering:
- CSV format (for spreadsheets)
- JSON format (full database export)
- SQLite database (single file, portable)
- GraphML format (Neo4j graph export)
- Qdrant vectors backup

Filtering options:
- By category (P, B, C, U)
- By manufacturer
- By translation status
- By date range

Usage:
    python scripts/export_data.py --all                    # Export all formats
    python scripts/export_data.py --csv                    # Export to CSV
    python scripts/export_data.py --sqlite                 # Export to SQLite
    python scripts/export_data.py --json                   # Export to JSON
    python scripts/export_data.py --graphml                # Export Neo4j to GraphML
    python scripts/export_data.py --qdrant                 # Export Qdrant vectors

    # Filtering:
    python scripts/export_data.py --category P             # Only powertrain codes
    python scripts/export_data.py --manufacturer Toyota    # Only Toyota codes
    python scripts/export_data.py --translated             # Only translated codes
    python scripts/export_data.py --untranslated           # Only untranslated codes
    python scripts/export_data.py --since 2024-01-01       # Since date
    python scripts/export_data.py --until 2024-12-31       # Until date
"""

import argparse
import csv
import gzip
import json
import logging
import sqlite3
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set
from xml.dom import minidom

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from backend.app.core.config import settings

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# Export directory
EXPORT_DIR = PROJECT_ROOT / "data" / "exports"
DATA_DIR = PROJECT_ROOT / "data"

# Category mapping
CATEGORY_MAP = {
    "P": "powertrain",
    "B": "body",
    "C": "chassis",
    "U": "network",
}


class ExportFilter:
    """Filter configuration for data export."""

    def __init__(
        self,
        categories: Optional[List[str]] = None,
        manufacturer: Optional[str] = None,
        translated_only: bool = False,
        untranslated_only: bool = False,
        since: Optional[datetime] = None,
        until: Optional[datetime] = None,
    ):
        self.categories = categories
        self.manufacturer = manufacturer
        self.translated_only = translated_only
        self.untranslated_only = untranslated_only
        self.since = since
        self.until = until

    def matches_dtc(self, dtc_data: Dict[str, Any]) -> bool:
        """Check if a DTC code matches the filter criteria."""
        # Category filter
        if self.categories:
            code = dtc_data.get("code", "")
            if code and code[0].upper() not in self.categories:
                return False

        # Manufacturer filter
        if self.manufacturer:
            applicable_makes = dtc_data.get("applicable_makes", [])
            manufacturer_field = dtc_data.get("manufacturer")
            if applicable_makes:
                if self.manufacturer.lower() not in [m.lower() for m in applicable_makes]:
                    return False
            elif manufacturer_field:
                if self.manufacturer.lower() != manufacturer_field.lower():
                    return False
            else:
                return False

        # Translation status filter
        if self.translated_only:
            if not dtc_data.get("description_hu"):
                return False

        if self.untranslated_only:
            if dtc_data.get("description_hu"):
                return False

        # Date range filter
        created_at = dtc_data.get("created_at")
        if created_at:
            if isinstance(created_at, str):
                try:
                    created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                except ValueError:
                    pass

            if isinstance(created_at, datetime):
                if self.since and created_at < self.since:
                    return False
                if self.until and created_at > self.until:
                    return False

        return True

    def matches_node(self, label: str, node_data: Dict[str, Any]) -> bool:
        """Check if a Neo4j node matches the filter criteria."""
        if label == "DTCNode":
            return self.matches_dtc(node_data)
        return True


def sanitize_for_export(data: Any) -> Any:
    """
    Sanitize data for export, removing sensitive information.

    Security: Ensures no passwords, tokens, or sensitive data are exported.
    """
    if isinstance(data, dict):
        sanitized = {}
        sensitive_keys = {
            "password", "hashed_password", "secret", "token", "api_key",
            "private_key", "credentials", "auth", "session_token",
        }
        for key, value in data.items():
            key_lower = key.lower()
            if any(s in key_lower for s in sensitive_keys):
                continue  # Skip sensitive fields
            sanitized[key] = sanitize_for_export(value)
        return sanitized
    elif isinstance(data, list):
        return [sanitize_for_export(item) for item in data]
    elif isinstance(data, bytes):
        return data.hex()
    elif hasattr(data, 'isoformat'):
        return data.isoformat()
    return data


def ensure_export_dir(subdirectory: Optional[str] = None) -> Path:
    """Create export directory with timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if subdirectory:
        export_path = EXPORT_DIR / subdirectory / timestamp
    else:
        export_path = EXPORT_DIR / timestamp
    export_path.mkdir(parents=True, exist_ok=True)
    logger.info(f"Export directory: {export_path}")
    return export_path


def get_postgres_data(export_filter: Optional[ExportFilter] = None) -> Dict[str, Any]:
    """Fetch all data from PostgreSQL with optional filtering."""
    from sqlalchemy import create_engine, inspect, text
    from sqlalchemy.orm import Session

    # Convert async URL to sync
    url = settings.DATABASE_URL
    if url.startswith("postgresql+asyncpg://"):
        url = url.replace("postgresql+asyncpg://", "postgresql://")

    engine = create_engine(url)
    inspector = inspect(engine)

    data: Dict[str, Any] = {}

    with Session(engine) as session:
        table_names = inspector.get_table_names()

        for table_name in table_names:
            # Skip alembic version table and sensitive tables
            if table_name in ("alembic_version", "users"):
                continue

            columns = [col["name"] for col in inspector.get_columns(table_name)]
            result = session.execute(text(f'SELECT * FROM "{table_name}"'))
            rows = result.fetchall()

            table_data = []
            for row in rows:
                row_dict = {}
                for i, col in enumerate(columns):
                    value = row[i]
                    row_dict[col] = sanitize_for_export(value)

                # Apply filter for dtc_codes table
                if table_name == "dtc_codes" and export_filter:
                    if not export_filter.matches_dtc(row_dict):
                        continue

                table_data.append(row_dict)

            data[table_name] = {
                "columns": columns,
                "data": table_data,
            }

    return data


def get_neo4j_data(export_filter: Optional[ExportFilter] = None) -> Dict[str, Any]:
    """Fetch all data from Neo4j with optional filtering."""
    try:
        from neo4j import GraphDatabase

        driver = GraphDatabase.driver(
            settings.NEO4J_URI,
            auth=(settings.NEO4J_USER, settings.NEO4J_PASSWORD)
        )

        data: Dict[str, Any] = {"nodes": {}, "relationships": []}

        with driver.session() as session:
            # Get labels
            labels_result = session.run("CALL db.labels()")
            labels = [record["label"] for record in labels_result]

            # Export nodes
            for label in labels:
                result = session.run(f"MATCH (n:{label}) RETURN n, elementId(n) as node_id")
                nodes = []
                for record in result:
                    node = record["n"]
                    node_data = sanitize_for_export(dict(node))
                    node_data["_element_id"] = record["node_id"]

                    # Apply filter
                    if export_filter and not export_filter.matches_node(label, node_data):
                        continue

                    nodes.append(node_data)
                data["nodes"][label] = nodes

            # Export relationships
            result = session.run("""
                MATCH (a)-[r]->(b)
                RETURN
                    labels(a)[0] as start_label,
                    elementId(a) as start_id,
                    properties(a) as start_props,
                    type(r) as rel_type,
                    properties(r) as rel_props,
                    labels(b)[0] as end_label,
                    elementId(b) as end_id,
                    properties(b) as end_props
            """)

            for record in result:
                rel_data = {
                    "start_label": record["start_label"],
                    "start_id": record["start_id"],
                    "start_props": sanitize_for_export(dict(record["start_props"]) if record["start_props"] else {}),
                    "type": record["rel_type"],
                    "properties": sanitize_for_export(dict(record["rel_props"]) if record["rel_props"] else {}),
                    "end_label": record["end_label"],
                    "end_id": record["end_id"],
                    "end_props": sanitize_for_export(dict(record["end_props"]) if record["end_props"] else {}),
                }
                data["relationships"].append(rel_data)

        driver.close()
        return data

    except Exception as e:
        logger.warning(f"Could not connect to Neo4j: {e}")
        return {"nodes": {}, "relationships": []}


def get_qdrant_data(include_vectors: bool = False) -> Dict[str, Any]:
    """Fetch all data from Qdrant collections."""
    try:
        from qdrant_client import QdrantClient

        if settings.QDRANT_URL:
            client = QdrantClient(
                url=settings.QDRANT_URL,
                api_key=settings.QDRANT_API_KEY,
            )
        else:
            client = QdrantClient(
                host=settings.QDRANT_HOST,
                port=settings.QDRANT_PORT,
            )

        data: Dict[str, Any] = {"collections": {}}

        collections = client.get_collections().collections
        for collection in collections:
            name = collection.name
            info = client.get_collection(name)

            collection_data = {
                "points_count": info.points_count,
                "vector_size": info.config.params.vectors.size if hasattr(info.config.params.vectors, 'size') else None,
                "distance": str(info.config.params.vectors.distance) if hasattr(info.config.params.vectors, 'distance') else None,
                "points": [],
            }

            # Scroll through all points
            offset = None
            while True:
                result = client.scroll(
                    collection_name=name,
                    offset=offset,
                    limit=100,
                    with_payload=True,
                    with_vectors=include_vectors,
                )

                points, next_offset = result

                for point in points:
                    point_data = {
                        "id": point.id,
                        "payload": sanitize_for_export(point.payload),
                    }
                    if include_vectors and point.vector:
                        point_data["vector"] = list(point.vector) if hasattr(point.vector, '__iter__') else point.vector
                    collection_data["points"].append(point_data)

                if next_offset is None:
                    break
                offset = next_offset

            data["collections"][name] = collection_data

        return data

    except Exception as e:
        logger.warning(f"Could not connect to Qdrant: {e}")
        return {"collections": {}}


def export_to_csv(
    export_path: Path,
    export_filter: Optional[ExportFilter] = None,
) -> bool:
    """
    Export data to CSV files.

    Creates one CSV file per table/collection.
    """
    logger.info("Starting CSV export...")

    try:
        csv_path = export_path / "csv"
        csv_path.mkdir(exist_ok=True)

        # Export PostgreSQL tables
        logger.info("  Exporting PostgreSQL tables to CSV...")
        try:
            pg_data = get_postgres_data(export_filter)
            for table_name, table_info in pg_data.items():
                if not table_info["data"]:
                    continue

                output_file = csv_path / f"pg_{table_name}.csv"
                with open(output_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=table_info["columns"])
                    writer.writeheader()
                    for row in table_info["data"]:
                        # Convert lists to JSON strings for CSV
                        row_copy = {}
                        for k, v in row.items():
                            if isinstance(v, (list, dict)):
                                row_copy[k] = json.dumps(v, ensure_ascii=False)
                            else:
                                row_copy[k] = v
                        writer.writerow(row_copy)
                logger.info(f"    Exported {table_name}: {len(table_info['data'])} rows")
        except Exception as e:
            logger.warning(f"  PostgreSQL CSV export failed: {e}")

        # Export Neo4j nodes
        logger.info("  Exporting Neo4j nodes to CSV...")
        try:
            neo4j_data = get_neo4j_data(export_filter)
            for label, nodes in neo4j_data["nodes"].items():
                if not nodes:
                    continue

                # Get all unique keys across all nodes
                all_keys = set()
                for node in nodes:
                    all_keys.update(node.keys())
                all_keys = sorted(all_keys)

                output_file = csv_path / f"neo4j_{label}.csv"
                with open(output_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=all_keys)
                    writer.writeheader()
                    for node in nodes:
                        row = {}
                        for key in all_keys:
                            value = node.get(key, "")
                            if isinstance(value, (list, dict)):
                                value = json.dumps(value, ensure_ascii=False)
                            row[key] = value
                        writer.writerow(row)
                logger.info(f"    Exported {label}: {len(nodes)} nodes")

            # Export relationships
            if neo4j_data["relationships"]:
                output_file = csv_path / "neo4j_relationships.csv"
                fieldnames = ["start_label", "start_id", "rel_type", "rel_props", "end_label", "end_id"]
                with open(output_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(fieldnames)
                    for rel in neo4j_data["relationships"]:
                        writer.writerow([
                            rel["start_label"],
                            rel["start_id"],
                            rel["type"],
                            json.dumps(rel["properties"], ensure_ascii=False),
                            rel["end_label"],
                            rel["end_id"],
                        ])
                logger.info(f"    Exported relationships: {len(neo4j_data['relationships'])}")
        except Exception as e:
            logger.warning(f"  Neo4j CSV export failed: {e}")

        # Export DTC codes from JSON files
        logger.info("  Exporting DTC codes from JSON files to CSV...")
        try:
            dtc_json_path = DATA_DIR / "dtc_codes" / "all_codes_merged.json"
            if dtc_json_path.exists():
                with open(dtc_json_path, 'r', encoding='utf-8') as f:
                    dtc_data = json.load(f)

                codes = dtc_data.get("codes", [])

                # Apply filter
                if export_filter:
                    codes = [c for c in codes if export_filter.matches_dtc(c)]

                if codes:
                    # Get all unique keys
                    all_keys = set()
                    for code in codes:
                        all_keys.update(code.keys())
                    all_keys = sorted(all_keys)

                    output_file = csv_path / "dtc_codes_full.csv"
                    with open(output_file, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.DictWriter(f, fieldnames=all_keys)
                        writer.writeheader()
                        for code in codes:
                            row = {}
                            for key in all_keys:
                                value = code.get(key, "")
                                if isinstance(value, (list, dict)):
                                    value = json.dumps(value, ensure_ascii=False)
                                row[key] = value
                            writer.writerow(row)
                    logger.info(f"    Exported dtc_codes_full: {len(codes)} codes")
        except Exception as e:
            logger.warning(f"  DTC JSON CSV export failed: {e}")

        logger.info(f"CSV export complete: {csv_path}")
        return True

    except Exception as e:
        logger.error(f"CSV export failed: {e}")
        return False


def export_to_sqlite(
    export_path: Path,
    export_filter: Optional[ExportFilter] = None,
) -> bool:
    """
    Export data to SQLite database.

    Creates a single portable SQLite file.
    """
    logger.info("Starting SQLite export...")

    try:
        output_file = export_path / "autocognitix_export.db"

        # Remove existing file
        if output_file.exists():
            output_file.unlink()

        conn = sqlite3.connect(output_file)
        cursor = conn.cursor()

        # Export PostgreSQL tables
        logger.info("  Exporting PostgreSQL tables to SQLite...")
        try:
            pg_data = get_postgres_data(export_filter)
            for table_name, table_info in pg_data.items():
                if not table_info["data"]:
                    continue

                columns = table_info["columns"]

                # Create table (infer types from first row)
                col_defs = []
                first_row = table_info["data"][0] if table_info["data"] else {}
                for col in columns:
                    value = first_row.get(col)
                    if isinstance(value, bool):
                        col_type = "INTEGER"
                    elif isinstance(value, int):
                        col_type = "INTEGER"
                    elif isinstance(value, float):
                        col_type = "REAL"
                    else:
                        col_type = "TEXT"
                    col_defs.append(f'"{col}" {col_type}')

                create_sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({", ".join(col_defs)})'
                cursor.execute(create_sql)

                # Insert data
                placeholders = ", ".join(["?" for _ in columns])
                quoted_columns = ", ".join([f'"{c}"' for c in columns])
                insert_sql = f'INSERT INTO "{table_name}" ({quoted_columns}) VALUES ({placeholders})'

                for row in table_info["data"]:
                    values = []
                    for col in columns:
                        value = row.get(col)
                        if isinstance(value, (list, dict)):
                            value = json.dumps(value, ensure_ascii=False)
                        elif isinstance(value, bool):
                            value = 1 if value else 0
                        values.append(value)
                    cursor.execute(insert_sql, values)

                logger.info(f"    Exported {table_name}: {len(table_info['data'])} rows")
        except Exception as e:
            logger.warning(f"  PostgreSQL SQLite export failed: {e}")

        # Export Neo4j nodes
        logger.info("  Exporting Neo4j nodes to SQLite...")
        try:
            neo4j_data = get_neo4j_data(export_filter)
            for label, nodes in neo4j_data["nodes"].items():
                if not nodes:
                    continue

                # Get all unique keys
                all_keys = set()
                for node in nodes:
                    all_keys.update(node.keys())
                all_keys = sorted(all_keys)

                # Create table
                col_defs = [f'"{key}" TEXT' for key in all_keys]
                table_name = f"neo4j_{label}"
                create_sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({", ".join(col_defs)})'
                cursor.execute(create_sql)

                # Insert data
                placeholders = ", ".join(["?" for _ in all_keys])
                quoted_keys = ", ".join([f'"{c}"' for c in all_keys])
                insert_sql = f'INSERT INTO "{table_name}" ({quoted_keys}) VALUES ({placeholders})'

                for node in nodes:
                    values = []
                    for key in all_keys:
                        value = node.get(key, "")
                        if isinstance(value, (list, dict)):
                            value = json.dumps(value, ensure_ascii=False)
                        values.append(value)
                    cursor.execute(insert_sql, values)

                logger.info(f"    Exported {label}: {len(nodes)} nodes")

            # Export relationships
            if neo4j_data["relationships"]:
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS "neo4j_relationships" (
                        start_label TEXT,
                        start_id TEXT,
                        rel_type TEXT,
                        rel_props TEXT,
                        end_label TEXT,
                        end_id TEXT
                    )
                ''')

                for rel in neo4j_data["relationships"]:
                    cursor.execute(
                        'INSERT INTO "neo4j_relationships" VALUES (?, ?, ?, ?, ?, ?)',
                        (
                            rel["start_label"],
                            rel["start_id"],
                            rel["type"],
                            json.dumps(rel["properties"], ensure_ascii=False),
                            rel["end_label"],
                            rel["end_id"],
                        )
                    )
                logger.info(f"    Exported relationships: {len(neo4j_data['relationships'])}")
        except Exception as e:
            logger.warning(f"  Neo4j SQLite export failed: {e}")

        # Create metadata table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS "_export_metadata" (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        ''')
        cursor.execute(
            'INSERT INTO "_export_metadata" VALUES (?, ?)',
            ("export_time", datetime.now().isoformat())
        )
        cursor.execute(
            'INSERT INTO "_export_metadata" VALUES (?, ?)',
            ("source", "AutoCognitix")
        )
        cursor.execute(
            'INSERT INTO "_export_metadata" VALUES (?, ?)',
            ("version", "2.0")
        )

        # Create indexes for better query performance
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_dtc_codes_code ON dtc_codes(code)')

        conn.commit()
        conn.close()

        file_size = output_file.stat().st_size / 1024
        logger.info(f"SQLite export complete: {output_file} ({file_size:.1f} KB)")
        return True

    except Exception as e:
        logger.error(f"SQLite export failed: {e}")
        return False


def export_to_json(
    export_path: Path,
    export_filter: Optional[ExportFilter] = None,
) -> bool:
    """
    Export all data to JSON format.

    Creates a single comprehensive JSON file.
    """
    logger.info("Starting JSON export...")

    try:
        export_data: Dict[str, Any] = {
            "export_time": datetime.now().isoformat(),
            "source": "AutoCognitix",
            "version": "2.0",
            "filter_applied": {
                "categories": export_filter.categories if export_filter else None,
                "manufacturer": export_filter.manufacturer if export_filter else None,
                "translated_only": export_filter.translated_only if export_filter else False,
                "untranslated_only": export_filter.untranslated_only if export_filter else False,
            },
            "postgresql": {},
            "neo4j": {},
            "qdrant": {},
            "translations": {},
        }

        # Export PostgreSQL
        logger.info("  Exporting PostgreSQL to JSON...")
        try:
            pg_data = get_postgres_data(export_filter)
            export_data["postgresql"] = {
                table: {
                    "row_count": len(info["data"]),
                    "columns": info["columns"],
                    "data": info["data"],
                }
                for table, info in pg_data.items()
            }
            total_rows = sum(len(info["data"]) for info in pg_data.values())
            logger.info(f"    PostgreSQL: {len(pg_data)} tables, {total_rows} total rows")
        except Exception as e:
            logger.warning(f"  PostgreSQL JSON export failed: {e}")

        # Export Neo4j
        logger.info("  Exporting Neo4j to JSON...")
        try:
            neo4j_data = get_neo4j_data(export_filter)
            export_data["neo4j"] = neo4j_data
            total_nodes = sum(len(nodes) for nodes in neo4j_data["nodes"].values())
            logger.info(f"    Neo4j: {len(neo4j_data['nodes'])} labels, {total_nodes} nodes, {len(neo4j_data['relationships'])} relationships")
        except Exception as e:
            logger.warning(f"  Neo4j JSON export failed: {e}")

        # Export Qdrant (payloads only, no vectors for standard export)
        logger.info("  Exporting Qdrant to JSON...")
        try:
            qdrant_data = get_qdrant_data(include_vectors=False)
            export_data["qdrant"] = qdrant_data
            for name, collection in qdrant_data.get("collections", {}).items():
                logger.info(f"    Qdrant {name}: {len(collection.get('points', []))} points")
        except Exception as e:
            logger.warning(f"  Qdrant JSON export failed: {e}")

        # Export translation cache
        logger.info("  Exporting translations...")
        try:
            translation_cache_path = DATA_DIR / "dtc_codes" / "translation_cache.json"
            if translation_cache_path.exists():
                with open(translation_cache_path, 'r', encoding='utf-8') as f:
                    export_data["translations"] = json.load(f)
                logger.info(f"    Translations: {len(export_data['translations'])} entries")
        except Exception as e:
            logger.warning(f"  Translations export failed: {e}")

        # Save JSON file
        output_file = export_path / "autocognitix_export.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)

        file_size = output_file.stat().st_size / 1024
        logger.info(f"JSON export complete: {output_file} ({file_size:.1f} KB)")
        return True

    except Exception as e:
        logger.error(f"JSON export failed: {e}")
        return False


def export_to_graphml(
    export_path: Path,
    export_filter: Optional[ExportFilter] = None,
) -> bool:
    """
    Export Neo4j graph to GraphML format.

    GraphML is an XML-based format that can be imported into:
    - Neo4j (via APOC)
    - Gephi
    - yEd
    - Cytoscape
    """
    logger.info("Starting GraphML export...")

    try:
        neo4j_data = get_neo4j_data(export_filter)

        # Create GraphML root element
        graphml = ET.Element("graphml")
        graphml.set("xmlns", "http://graphml.graphdrawing.org/xmlns")
        graphml.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
        graphml.set("xsi:schemaLocation",
                    "http://graphml.graphdrawing.org/xmlns http://graphml.graphdrawing.org/xmlns/1.0/graphml.xsd")

        # Collect all unique property keys from nodes and relationships
        node_keys: Dict[str, Set[str]] = {}  # label -> set of property names
        rel_keys: Set[str] = set()

        for label, nodes in neo4j_data["nodes"].items():
            node_keys[label] = set()
            for node in nodes:
                for key in node.keys():
                    if not key.startswith("_"):
                        node_keys[label].add(key)

        for rel in neo4j_data["relationships"]:
            for key in rel.get("properties", {}).keys():
                rel_keys.add(key)

        # Define key elements for node properties
        all_node_keys = set()
        for keys in node_keys.values():
            all_node_keys.update(keys)

        for key_name in sorted(all_node_keys):
            key_elem = ET.SubElement(graphml, "key")
            key_elem.set("id", f"n_{key_name}")
            key_elem.set("for", "node")
            key_elem.set("attr.name", key_name)
            key_elem.set("attr.type", "string")

        # Add label key for nodes
        label_key = ET.SubElement(graphml, "key")
        label_key.set("id", "n_label")
        label_key.set("for", "node")
        label_key.set("attr.name", "label")
        label_key.set("attr.type", "string")

        # Define key elements for edge properties
        for key_name in sorted(rel_keys):
            key_elem = ET.SubElement(graphml, "key")
            key_elem.set("id", f"e_{key_name}")
            key_elem.set("for", "edge")
            key_elem.set("attr.name", key_name)
            key_elem.set("attr.type", "string")

        # Add relationship type key
        rel_type_key = ET.SubElement(graphml, "key")
        rel_type_key.set("id", "e_type")
        rel_type_key.set("for", "edge")
        rel_type_key.set("attr.name", "type")
        rel_type_key.set("attr.type", "string")

        # Create graph element
        graph = ET.SubElement(graphml, "graph")
        graph.set("id", "G")
        graph.set("edgedefault", "directed")

        # Track node IDs for relationship creation
        node_id_map: Dict[str, str] = {}  # element_id -> graph_id
        node_counter = 0

        # Add nodes
        for label, nodes in neo4j_data["nodes"].items():
            for node in nodes:
                node_id = f"n{node_counter}"
                node_counter += 1

                element_id = node.get("_element_id", "")
                if element_id:
                    node_id_map[element_id] = node_id

                node_elem = ET.SubElement(graph, "node")
                node_elem.set("id", node_id)

                # Add label
                label_data = ET.SubElement(node_elem, "data")
                label_data.set("key", "n_label")
                label_data.text = label

                # Add properties
                for key, value in node.items():
                    if key.startswith("_"):
                        continue
                    data_elem = ET.SubElement(node_elem, "data")
                    data_elem.set("key", f"n_{key}")
                    if isinstance(value, (list, dict)):
                        data_elem.text = json.dumps(value, ensure_ascii=False)
                    else:
                        data_elem.text = str(value) if value is not None else ""

        logger.info(f"  Added {node_counter} nodes")

        # Add edges (relationships)
        edge_counter = 0
        for rel in neo4j_data["relationships"]:
            source_id = node_id_map.get(rel["start_id"])
            target_id = node_id_map.get(rel["end_id"])

            if not source_id or not target_id:
                continue

            edge_id = f"e{edge_counter}"
            edge_counter += 1

            edge_elem = ET.SubElement(graph, "edge")
            edge_elem.set("id", edge_id)
            edge_elem.set("source", source_id)
            edge_elem.set("target", target_id)

            # Add relationship type
            type_data = ET.SubElement(edge_elem, "data")
            type_data.set("key", "e_type")
            type_data.text = rel["type"]

            # Add properties
            for key, value in rel.get("properties", {}).items():
                data_elem = ET.SubElement(edge_elem, "data")
                data_elem.set("key", f"e_{key}")
                if isinstance(value, (list, dict)):
                    data_elem.text = json.dumps(value, ensure_ascii=False)
                else:
                    data_elem.text = str(value) if value is not None else ""

        logger.info(f"  Added {edge_counter} edges")

        # Write to file with pretty printing
        output_file = export_path / "neo4j_graph.graphml"

        # Pretty print XML
        rough_string = ET.tostring(graphml, encoding='unicode')
        reparsed = minidom.parseString(rough_string)
        pretty_xml = reparsed.toprettyxml(indent="  ")

        # Remove extra blank lines
        lines = [line for line in pretty_xml.split('\n') if line.strip()]
        pretty_xml = '\n'.join(lines)

        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(pretty_xml)

        file_size = output_file.stat().st_size / 1024
        logger.info(f"GraphML export complete: {output_file} ({file_size:.1f} KB)")
        return True

    except Exception as e:
        logger.error(f"GraphML export failed: {e}")
        return False


def export_qdrant_vectors(export_path: Path) -> bool:
    """
    Export Qdrant vectors for backup purposes.

    Includes full vectors for restore capability.
    """
    logger.info("Starting Qdrant vectors export...")

    try:
        qdrant_data = get_qdrant_data(include_vectors=True)

        if not qdrant_data.get("collections"):
            logger.warning("No Qdrant collections found")
            return True

        # Save as compressed JSON (vectors are large)
        output_file = export_path / "qdrant_vectors.json.gz"
        with gzip.open(output_file, 'wt', encoding='utf-8') as f:
            json.dump(qdrant_data, f, ensure_ascii=False, default=str)

        file_size = output_file.stat().st_size / 1024
        total_points = sum(
            len(c.get("points", []))
            for c in qdrant_data.get("collections", {}).values()
        )
        logger.info(f"Qdrant vectors export complete: {output_file} ({file_size:.1f} KB, {total_points} points)")
        return True

    except Exception as e:
        logger.error(f"Qdrant vectors export failed: {e}")
        return False


def export_to_excel(
    export_path: Path,
    export_filter: Optional[ExportFilter] = None,
) -> bool:
    """
    Export data to Excel format (.xlsx).

    Creates a multi-sheet Excel workbook with:
    - DTC codes (with categories in separate sheets)
    - Diagnosis history
    - Neo4j graph data (nodes by label)
    - Translation summary

    Requires openpyxl: pip install openpyxl
    """
    logger.info("Starting Excel export...")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        return False

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Export Summary"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def style_header(ws, row: int = 1) -> None:
            """Apply header styles to the first row."""
            for cell in ws[row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        def auto_column_width(ws) -> None:
            """Auto-adjust column widths based on content."""
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width

        # Summary sheet
        ws["A1"] = "AutoCognitix Data Export"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A3"] = "Export Time:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A4"] = "Filter Applied:"
        filter_desc = "None"
        if export_filter:
            filter_parts = []
            if export_filter.categories:
                filter_parts.append(f"Categories: {','.join(export_filter.categories)}")
            if export_filter.manufacturer:
                filter_parts.append(f"Manufacturer: {export_filter.manufacturer}")
            if export_filter.translated_only:
                filter_parts.append("Translated only")
            if export_filter.untranslated_only:
                filter_parts.append("Untranslated only")
            filter_desc = "; ".join(filter_parts) if filter_parts else "None"
        ws["B4"] = filter_desc

        # Export PostgreSQL tables
        logger.info("  Exporting PostgreSQL tables to Excel...")
        try:
            pg_data = get_postgres_data(export_filter)
            for table_name, table_info in pg_data.items():
                if not table_info["data"]:
                    continue

                # Create sheet for table
                sheet_name = table_name[:31]  # Excel sheet name limit
                ws_table = wb.create_sheet(title=sheet_name)

                # Headers
                columns = table_info["columns"]
                for col_idx, col_name in enumerate(columns, start=1):
                    ws_table.cell(row=1, column=col_idx, value=col_name)
                style_header(ws_table)

                # Data rows
                for row_idx, row_data in enumerate(table_info["data"], start=2):
                    for col_idx, col_name in enumerate(columns, start=1):
                        value = row_data.get(col_name, "")
                        if isinstance(value, (list, dict)):
                            value = json.dumps(value, ensure_ascii=False)
                        ws_table.cell(row=row_idx, column=col_idx, value=value)

                auto_column_width(ws_table)
                logger.info(f"    Exported {table_name}: {len(table_info['data'])} rows")
        except Exception as e:
            logger.warning(f"  PostgreSQL Excel export failed: {e}")

        # Export Neo4j nodes
        logger.info("  Exporting Neo4j nodes to Excel...")
        try:
            neo4j_data = get_neo4j_data(export_filter)
            for label, nodes in neo4j_data["nodes"].items():
                if not nodes:
                    continue

                # Create sheet for node type
                sheet_name = f"Neo4j_{label}"[:31]
                ws_neo4j = wb.create_sheet(title=sheet_name)

                # Get all unique keys
                all_keys = set()
                for node in nodes:
                    all_keys.update(k for k in node.keys() if not k.startswith("_"))
                all_keys = sorted(all_keys)

                # Headers
                for col_idx, key in enumerate(all_keys, start=1):
                    ws_neo4j.cell(row=1, column=col_idx, value=key)
                style_header(ws_neo4j)

                # Data rows
                for row_idx, node in enumerate(nodes, start=2):
                    for col_idx, key in enumerate(all_keys, start=1):
                        value = node.get(key, "")
                        if isinstance(value, (list, dict)):
                            value = json.dumps(value, ensure_ascii=False)
                        ws_neo4j.cell(row=row_idx, column=col_idx, value=value)

                auto_column_width(ws_neo4j)
                logger.info(f"    Exported {label}: {len(nodes)} nodes")
        except Exception as e:
            logger.warning(f"  Neo4j Excel export failed: {e}")

        # Export DTC codes from JSON files with category breakdown
        logger.info("  Exporting DTC codes from JSON files to Excel...")
        try:
            dtc_json_path = DATA_DIR / "dtc_codes" / "all_codes_merged.json"
            if dtc_json_path.exists():
                with open(dtc_json_path, 'r', encoding='utf-8') as f:
                    dtc_data = json.load(f)

                codes = dtc_data.get("codes", [])

                # Apply filter
                if export_filter:
                    codes = [c for c in codes if export_filter.matches_dtc(c)]

                # Group by category prefix
                categories = {"P": [], "B": [], "C": [], "U": [], "Other": []}
                for code in codes:
                    code_str = code.get("code", "")
                    if code_str:
                        prefix = code_str[0].upper()
                        if prefix in categories:
                            categories[prefix].append(code)
                        else:
                            categories["Other"].append(code)

                # Create sheet for all codes
                ws_all = wb.create_sheet(title="All_DTC_Codes")

                if codes:
                    # Get all unique keys
                    all_keys = set()
                    for code in codes:
                        all_keys.update(code.keys())
                    # Order common keys first
                    priority_keys = ["code", "description_en", "description_hu", "category", "severity"]
                    ordered_keys = [k for k in priority_keys if k in all_keys]
                    ordered_keys.extend(k for k in sorted(all_keys) if k not in priority_keys)

                    # Headers
                    for col_idx, key in enumerate(ordered_keys, start=1):
                        ws_all.cell(row=1, column=col_idx, value=key)
                    style_header(ws_all)

                    # Data rows
                    for row_idx, code in enumerate(codes, start=2):
                        for col_idx, key in enumerate(ordered_keys, start=1):
                            value = code.get(key, "")
                            if isinstance(value, (list, dict)):
                                value = json.dumps(value, ensure_ascii=False)
                            ws_all.cell(row=row_idx, column=col_idx, value=value)

                    auto_column_width(ws_all)
                    logger.info(f"    Exported All_DTC_Codes: {len(codes)} codes")

                # Create sheets per category
                category_names = {
                    "P": "Powertrain_DTC",
                    "B": "Body_DTC",
                    "C": "Chassis_DTC",
                    "U": "Network_DTC",
                    "Other": "Other_DTC",
                }

                for prefix, cat_codes in categories.items():
                    if not cat_codes:
                        continue

                    sheet_name = category_names.get(prefix, f"{prefix}_DTC")[:31]
                    ws_cat = wb.create_sheet(title=sheet_name)

                    # Get keys for this category
                    cat_keys = set()
                    for code in cat_codes:
                        cat_keys.update(code.keys())
                    ordered_keys = [k for k in priority_keys if k in cat_keys]
                    ordered_keys.extend(k for k in sorted(cat_keys) if k not in priority_keys)

                    # Headers
                    for col_idx, key in enumerate(ordered_keys, start=1):
                        ws_cat.cell(row=1, column=col_idx, value=key)
                    style_header(ws_cat)

                    # Data rows
                    for row_idx, code in enumerate(cat_codes, start=2):
                        for col_idx, key in enumerate(ordered_keys, start=1):
                            value = code.get(key, "")
                            if isinstance(value, (list, dict)):
                                value = json.dumps(value, ensure_ascii=False)
                            ws_cat.cell(row=row_idx, column=col_idx, value=value)

                    auto_column_width(ws_cat)
                    logger.info(f"    Exported {sheet_name}: {len(cat_codes)} codes")

        except Exception as e:
            logger.warning(f"  DTC JSON Excel export failed: {e}")

        # Export translations summary
        logger.info("  Exporting translations summary to Excel...")
        try:
            translation_cache_path = DATA_DIR / "dtc_codes" / "translation_cache.json"
            if translation_cache_path.exists():
                with open(translation_cache_path, 'r', encoding='utf-8') as f:
                    translations = json.load(f)

                ws_trans = wb.create_sheet(title="Translations")
                ws_trans.cell(row=1, column=1, value="English Text")
                ws_trans.cell(row=1, column=2, value="Hungarian Translation")
                ws_trans.cell(row=1, column=3, value="Hash Key")
                style_header(ws_trans)

                for row_idx, (key, value) in enumerate(translations.items(), start=2):
                    if isinstance(value, dict):
                        ws_trans.cell(row=row_idx, column=1, value=value.get("original", ""))
                        ws_trans.cell(row=row_idx, column=2, value=value.get("translation", ""))
                    else:
                        ws_trans.cell(row=row_idx, column=2, value=value)
                    ws_trans.cell(row=row_idx, column=3, value=key)

                auto_column_width(ws_trans)
                logger.info(f"    Exported translations: {len(translations)} entries")
        except Exception as e:
            logger.warning(f"  Translations Excel export failed: {e}")

        # Update summary sheet with counts
        ws = wb["Export Summary"]
        ws["A6"] = "Sheet Summary:"
        ws["A6"].font = Font(bold=True)
        row = 7
        for sheet in wb.sheetnames:
            if sheet == "Export Summary":
                continue
            ws_sheet = wb[sheet]
            row_count = ws_sheet.max_row - 1  # Exclude header
            ws[f"A{row}"] = sheet
            ws[f"B{row}"] = f"{row_count} rows"
            row += 1

        # Save workbook
        output_file = export_path / "autocognitix_export.xlsx"
        wb.save(output_file)

        file_size = output_file.stat().st_size / 1024
        logger.info(f"Excel export complete: {output_file} ({file_size:.1f} KB)")
        return True

    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        return False


def export_diagnosis_history(
    export_path: Path,
    export_filter: Optional[ExportFilter] = None,
) -> bool:
    """
    Export diagnosis history to JSON format.

    Includes all diagnosis sessions with their inputs and results.
    """
    logger.info("Starting diagnosis history export...")

    try:
        # Get diagnosis history from PostgreSQL
        from sqlalchemy import create_engine, text
        from sqlalchemy.orm import Session

        url = settings.DATABASE_URL
        if url.startswith("postgresql+asyncpg://"):
            url = url.replace("postgresql+asyncpg://", "postgresql://")

        engine = create_engine(url)

        history_data: Dict[str, Any] = {
            "export_time": datetime.now().isoformat(),
            "sessions": [],
        }

        with Session(engine) as session:
            # Query diagnosis sessions (table may not exist yet)
            try:
                result = session.execute(text("""
                    SELECT * FROM diagnosis_sessions
                    ORDER BY created_at DESC
                """))
                columns = result.keys()
                for row in result.fetchall():
                    session_data = dict(zip(columns, row))
                    history_data["sessions"].append(sanitize_for_export(session_data))

            except Exception as e:
                logger.warning(f"  Could not query diagnosis_sessions: {e}")

                # Try querying diagnosis_results if sessions table doesn't exist
                try:
                    result = session.execute(text("""
                        SELECT * FROM diagnosis_results
                        ORDER BY created_at DESC
                    """))
                    columns = result.keys()
                    for row in result.fetchall():
                        session_data = dict(zip(columns, row))
                        history_data["sessions"].append(sanitize_for_export(session_data))
                except Exception as e2:
                    logger.warning(f"  Could not query diagnosis_results: {e2}")

        # Save as JSON
        output_file = export_path / "diagnosis_history.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(history_data, f, indent=2, ensure_ascii=False, default=str)

        file_size = output_file.stat().st_size / 1024
        logger.info(f"Diagnosis history export complete: {output_file} ({file_size:.1f} KB, {len(history_data['sessions'])} sessions)")
        return True

    except Exception as e:
        logger.error(f"Diagnosis history export failed: {e}")
        return False


def export_translations_only(export_path: Path) -> bool:
    """
    Export only the translation cache as a standalone JSON file.

    Useful for translation management and backup.
    """
    logger.info("Starting translations export...")

    try:
        translation_cache_path = DATA_DIR / "dtc_codes" / "translation_cache.json"

        if not translation_cache_path.exists():
            logger.warning("Translation cache not found")
            return True

        with open(translation_cache_path, 'r', encoding='utf-8') as f:
            translations = json.load(f)

        # Export with metadata
        export_data = {
            "export_time": datetime.now().isoformat(),
            "source": "AutoCognitix Translation Cache",
            "total_translations": len(translations),
            "translations": translations,
        }

        output_file = export_path / "translations.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)

        file_size = output_file.stat().st_size / 1024
        logger.info(f"Translations export complete: {output_file} ({file_size:.1f} KB, {len(translations)} entries)")
        return True

    except Exception as e:
        logger.error(f"Translations export failed: {e}")
        return False


def create_export_manifest(
    export_path: Path,
    results: Dict[str, bool],
    export_filter: Optional[ExportFilter] = None,
) -> None:
    """Create manifest file with export details."""
    manifest = {
        "export_time": datetime.now().isoformat(),
        "export_path": str(export_path),
        "version": "2.0",
        "results": results,
        "filter": {
            "categories": export_filter.categories if export_filter else None,
            "manufacturer": export_filter.manufacturer if export_filter else None,
            "translated_only": export_filter.translated_only if export_filter else False,
            "untranslated_only": export_filter.untranslated_only if export_filter else False,
            "since": export_filter.since.isoformat() if export_filter and export_filter.since else None,
            "until": export_filter.until.isoformat() if export_filter and export_filter.until else None,
        },
        "files": [],
    }

    # List all files recursively
    for file in export_path.rglob("*"):
        if file.is_file() and file.name != "manifest.json":
            manifest["files"].append({
                "path": str(file.relative_to(export_path)),
                "size_kb": round(file.stat().st_size / 1024, 1),
                "format": file.suffix.lstrip("."),
            })

    manifest_file = export_path / "manifest.json"
    with open(manifest_file, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, indent=2)

    logger.info(f"Manifest created: {manifest_file}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Export AutoCognitix data to portable formats"
    )

    # Format options
    format_group = parser.add_argument_group("Export Formats")
    format_group.add_argument(
        "--csv",
        action="store_true",
        help="Export to CSV format",
    )
    format_group.add_argument(
        "--excel",
        action="store_true",
        help="Export to Excel format (.xlsx) with multiple sheets",
    )
    format_group.add_argument(
        "--sqlite",
        action="store_true",
        help="Export to SQLite database",
    )
    format_group.add_argument(
        "--json",
        action="store_true",
        help="Export to JSON format",
    )
    format_group.add_argument(
        "--graphml",
        action="store_true",
        help="Export Neo4j graph to GraphML format",
    )
    format_group.add_argument(
        "--qdrant",
        action="store_true",
        help="Export Qdrant vectors (with full vectors for backup)",
    )
    format_group.add_argument(
        "--translations",
        action="store_true",
        help="Export translations only",
    )
    format_group.add_argument(
        "--history",
        action="store_true",
        help="Export diagnosis history",
    )
    format_group.add_argument(
        "--all",
        action="store_true",
        help="Export to all formats",
    )

    # Filter options
    filter_group = parser.add_argument_group("Filtering Options")
    filter_group.add_argument(
        "--category",
        type=str,
        help="Filter by DTC category (P=powertrain, B=body, C=chassis, U=network). Can be comma-separated.",
    )
    filter_group.add_argument(
        "--manufacturer",
        type=str,
        help="Filter by manufacturer (e.g., Toyota, BMW)",
    )
    filter_group.add_argument(
        "--translated",
        action="store_true",
        help="Export only translated codes (with Hungarian description)",
    )
    filter_group.add_argument(
        "--untranslated",
        action="store_true",
        help="Export only untranslated codes (without Hungarian description)",
    )
    filter_group.add_argument(
        "--since",
        type=str,
        help="Export only data created since date (YYYY-MM-DD)",
    )
    filter_group.add_argument(
        "--until",
        type=str,
        help="Export only data created until date (YYYY-MM-DD)",
    )

    # Other options
    parser.add_argument(
        "--output-dir",
        type=str,
        help="Custom output directory (default: data/exports/<timestamp>)",
    )
    parser.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        help="Enable verbose logging",
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Build export filter
    export_filter = None
    if any([args.category, args.manufacturer, args.translated, args.untranslated, args.since, args.until]):
        categories = None
        if args.category:
            categories = [c.upper() for c in args.category.split(",")]

        since = None
        if args.since:
            since = datetime.fromisoformat(args.since)

        until = None
        if args.until:
            until = datetime.fromisoformat(args.until)

        export_filter = ExportFilter(
            categories=categories,
            manufacturer=args.manufacturer,
            translated_only=args.translated,
            untranslated_only=args.untranslated,
            since=since,
            until=until,
        )

        logger.info("Export filter applied:")
        if categories:
            logger.info(f"  Categories: {categories}")
        if args.manufacturer:
            logger.info(f"  Manufacturer: {args.manufacturer}")
        if args.translated:
            logger.info("  Translated only: Yes")
        if args.untranslated:
            logger.info("  Untranslated only: Yes")
        if since:
            logger.info(f"  Since: {since}")
        if until:
            logger.info(f"  Until: {until}")

    # Default to --all if no specific format is selected
    if not any([args.csv, args.excel, args.sqlite, args.json, args.graphml, args.qdrant, args.translations, args.history, args.all]):
        args.all = True

    # Create export directory
    if args.output_dir:
        export_path = Path(args.output_dir)
        export_path.mkdir(parents=True, exist_ok=True)
    else:
        export_path = ensure_export_dir()

    results: Dict[str, bool] = {}

    try:
        if args.csv or args.all:
            results["csv"] = export_to_csv(export_path, export_filter)

        if args.excel or args.all:
            results["excel"] = export_to_excel(export_path, export_filter)

        if args.sqlite or args.all:
            results["sqlite"] = export_to_sqlite(export_path, export_filter)

        if args.json or args.all:
            results["json"] = export_to_json(export_path, export_filter)

        if args.graphml or args.all:
            results["graphml"] = export_to_graphml(export_path, export_filter)

        if args.qdrant or args.all:
            results["qdrant"] = export_qdrant_vectors(export_path)

        if args.translations or args.all:
            results["translations"] = export_translations_only(export_path)

        if args.history or args.all:
            results["history"] = export_diagnosis_history(export_path, export_filter)

        # Create manifest
        create_export_manifest(export_path, results, export_filter)

        # Summary
        logger.info("=" * 50)
        logger.info("EXPORT SUMMARY")
        logger.info("=" * 50)
        for name, success in results.items():
            status = "SUCCESS" if success else "FAILED"
            logger.info(f"  {name}: {status}")
        logger.info(f"Export location: {export_path}")
        logger.info("=" * 50)

        # Exit with error if any export failed
        if not all(results.values()):
            sys.exit(1)

    except Exception as e:
        logger.error(f"Export failed: {e}")
        raise


if __name__ == "__main__":
    main()
